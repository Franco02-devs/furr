from django.shortcuts import render, redirect
from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from .forms import CustomUserCreationForm, CollaboratorCreationForm, InputRecordForm, OutputRecordForm
from .models import CustomUser, Collaborator, Record, AttendanceRecord
from .scripts import (
    getCollaboratorsActive,
    generateUniqueUsername,
    corregirFecha,
    eliminar,
    getAllCollaborators,
    calculateHoursWorked,
    getAllAttendanceRecordsT,
    getAllAttendanceRecordsTRange,
    getAllHoursWorked,
    isObserved,
    float_to_hms
)
from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.timezone import localtime
import pytz

## VIEWS
### LOGIN
def loginView(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect("home")
            else:
                messages.error(request, 'Nombre de usuario o contraseña incorrectos')
    else:
        form = AuthenticationForm() 

    return render(request, 'login.html', {'form': form})

### LOGOUT
def logoutView(request):
    logout(request)
    return redirect('login')

### HOME
def homeView(request):
    if request.user.is_authenticated:
        collaborators=getCollaboratorsActive()
        return render(request, 'home.html',{"collaborators":collaborators})
    return redirect('login')


###ADMIN
@login_required
def dashboardView(request):
    if (not request.user.is_staff and not request.user.isAdmin):
        return redirect('home')

    totalUsers = CustomUser.objects.count()
    totalCollaborators = Collaborator.objects.enables().count()
    totalAttendance = AttendanceRecord.objects.enables().count()

    context = {
        "total_users": totalUsers,
        "total_collaborators": totalCollaborators,
        "total_attendance": totalAttendance,    
    }
    return render(request, "adminDashboard.html", context)

### CREAR USUARIOS Y COLABORADORES
@user_passes_test(lambda u: u.is_superuser)
def createUserView(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password1'])
            user.save()
            collaborator=Collaborator.objects.create(
                user=user,
                favoritePlaceIsOffice=form.cleaned_data['favoritePlaceIsOffice']
            )
            collaborator.save()
            return redirect('user', user_id=user.id)
    else:
        form = CustomUserCreationForm()

    return render(request, 'createUser.html', {'form': form})
 
@user_passes_test(lambda u: u.is_superuser)
def userDetailView(request, user_id):
    user = CustomUser.objects.only('username', 'first_name', 'isAdmin', 'date_joined').get(id=user_id)
    return render(request, 'user.html', {'user': user})

@user_passes_test(lambda u: u.is_superuser or (u.isAdmin))
def createCollaboratorview(request):
    if request.method == 'POST':
        form = CollaboratorCreationForm(request.POST)
        if form.is_valid():
            collaborator = form.save(commit=False)
            password = form.cleaned_data['password1']
            name=form.cleaned_data['name']
            username=generateUniqueUsername(name)
            user = CustomUser.objects.create_user(
                username=username,
                password=password,
                first_name=name
            )
            user.username=username+str(user.id)+str(int(user.is_staff))
            user.save()
            collaborator.user = user
            collaborator.save()
            return redirect('collaborator', collaborator_id=collaborator.id)
    else:
        form = CollaboratorCreationForm()

    return render(request, 'createCollaborator.html', {'form': form})

@user_passes_test(lambda u: u.is_superuser or (u.isAdmin))
def collaboratorDetailView(request, collaborator_id):
    collaborator = Collaborator.objects.only('user__username', 'user__first_name', 'favoritePlaceIsOffice', 'isWorking').get(id=collaborator_id)
    return render(request, 'collaborator.html', {'collaborator': collaborator})

@login_required
def registerInputView(request):
    if request.method == 'POST':
        form = InputRecordForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            input = form.save(commit=False)
            collaborator = form.cleaned_data.get('collaborator')
            input.unTimelyDateTime=form.cleaned_data.get('unTimelyDateTime')
            input.collaborator = collaborator
            if collaborator.isWorking==-1:
                input.save()
                completeRecord=AttendanceRecord.objects.create(
                    inRecord=input,
                    collaborator=collaborator,
                )
                completeRecord.save()
                idCompleteRecord=completeRecord.id
                collaborator.isWorking=int(idCompleteRecord)
                collaborator.save()
                return redirect('record',record_id=input.id)
            else:
                messages.error(request, "Usted ya se encuentra trabajando, para ingresar de nuevo primero debe marcar su salida.")
                return render(request, 'registerInput.html', {'form': form})
    else:
        form = InputRecordForm(user=request.user)
    
    return render(request, 'registerInput.html', {'form': form})

@login_required
def registerOutputView(request):
    if request.method == 'POST':
        form = OutputRecordForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            output = form.save(commit=False)
            collaborator = form.cleaned_data.get('collaborator')
            output.collaborator = collaborator
            if collaborator.isWorking!=-1:
                output.typeIsInput=False
                completeAttendancePartial=AttendanceRecord.objects.get(id=collaborator.isWorking)
                output.placeIsOffice=completeAttendancePartial.inRecord.placeIsOffice
                output.save()
                completeAttendancePartial.outRecord=output
                completeAttendancePartial.save()
                collaborator.isWorking=-1
                collaborator.save()
                return redirect('attendanceRecord',attendanceRecord_id=completeAttendancePartial.id)
            else:
                messages.error(request, "Usted no esta trabajado, no puede marcar salida.")
                return render(request, 'registerOutput.html', {'form': form})
    else:
        form = OutputRecordForm(user=request.user)
    
    return render(request, 'registerOutput.html', {'form': form})

@login_required
def recordDetailView(request, record_id):
    record = Record.objects.get(id=record_id)
    return render(request, 'record.html', {'record': record})

@login_required
def attendanceRecordDetailView(request, attendanceRecord_id):
    attendanceRecord = AttendanceRecord.objects.get(id=attendanceRecord_id)
    if attendanceRecord.outRecord:
        worked_hours, _ =calculateHoursWorked(attendanceRecord)
    else:
        worked_hours = None
    context = {
        'attendanceRecord': attendanceRecord,
        'worked_hours': worked_hours
    }
    if (request.method == "POST" and "corregirEntrada" in request.POST) and (request.user.isAdmin or request.user.is_staff):
        corregirFecha(registro=attendanceRecord.inRecord)
        return render(request, 'attendanceRecord.html', context)
    if (request.method == "POST" and "corregirSalida" in request.POST) and (request.user.isAdmin or request.user.is_staff):
        corregirFecha(registro=attendanceRecord.outRecord)
        return render(request, 'attendanceRecord.html', context)
    if (request.method == "POST" and "corregirRegistro" in request.POST) and (request.user.isAdmin or request.user.is_staff):
        eliminar(registro=attendanceRecord)
        return redirect('adminDashboard')

    return render(request, 'attendanceRecord.html', context)

@login_required
def myRecords(request):
    if request.user.is_staff or request.user.isAdmin:
        collaborators = getAllCollaborators()
        selected_collaborator_id = request.GET.get('collaborator', None)
        if selected_collaborator_id:
            collaboratorS=Collaborator.objects.enables().get(id=selected_collaborator_id)
            records = getAllAttendanceRecordsT(collaboratorS,limit=10)
        else:
            records = AttendanceRecord.objects.none()
    else:
        collaborators = None
        records = getAllAttendanceRecordsT(collaborator=request.user.collaborator,limit=5)
        getAllHoursWorked(records)
        selected_collaborator_id=None
    totalHours,finalRecords=getAllHoursWorked(records)

    return render(request, 'myRecords.html', {
        'totalHours':totalHours,
        'records': finalRecords,
        'collaborators': collaborators,
        'selected_collaborator_id': selected_collaborator_id
    })
    
@login_required
def hoursWorked(request):
    user = request.user
    today = datetime.now().date()
    period = request.GET.get('period', 'week')
    
    if period == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif period == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif period == 'custom':
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            start_date = today
            end_date = today
    else:
        start_date = today
        end_date = today
    
    if user.is_staff or user.isAdmin:
        collaborators = getAllCollaborators()
        selected_collaborator_id = request.GET.get('collaborator', None)
        if selected_collaborator_id:
            collaboratorS=Collaborator.objects.enables().get(id=selected_collaborator_id)
            records=getAllAttendanceRecordsTRange(collaboratorS,start_date,end_date)
        else:
            records = AttendanceRecord.objects.none()
    else:
        collaborators = None
        selected_collaborator_id = None
        records=getAllAttendanceRecordsTRange(user.collaborator,start_date,end_date)

    total_hours,records=getAllHoursWorked(records)
    return render(request, 'hoursWorked.html', {
        'records': records,
        'collaborators': collaborators,
        'selected_collaborator_id': selected_collaborator_id,
        'total_hours': total_hours,
        'period': period,
        'start_date': start_date.strftime("%Y-%m-%d"),
        'end_date': end_date.strftime("%Y-%m-%d")
    })
    
@user_passes_test(lambda u: u.is_superuser or u.isAdmin)
def attendanceChatView(request):
    records = Record.objects.filter(
        dateTime__date=datetime.now().date()
    ).order_by("dateTime")
    records = records[::-1]
    return render(request, "attendanceChat.html", {"records": records})

def attendanceObservationsView(request):
    attendance_records = AttendanceRecord.objects.enables()

    observed_recordsUnTimely = []
    observedRecordsShort = []
    observedRecordsLarge = []
    totalUnTimely, totalShort, totalLarge = 0, 0, 0

    for record in attendance_records:
        observed_status = isObserved(record)
        hoursWorked, _  = calculateHoursWorked(record)
        
        if observed_status == 0:
            observed_recordsUnTimely.append((record, hoursWorked))
            totalUnTimely += hoursWorked
        elif observed_status == 1:
            observedRecordsShort.append((record, hoursWorked))
            totalShort += hoursWorked
        elif observed_status == 2:
            observedRecordsLarge.append((record, hoursWorked))
            totalLarge += hoursWorked

    return render(request, 'attendanceObservations.html', {
        'records_unTimely': observed_recordsUnTimely,
        'records_short': observedRecordsShort,
        'records_large': observedRecordsLarge,
        'total_unTimely': totalUnTimely,
        'total_short': totalShort,
        'total_large': totalLarge,
    })
def reporte_asistencia_template(request):
    current_year = datetime.now().year
    years = list(range(2024, current_year + 1))
    return render(request, "attendance_report.html", {"years": years})

def generar_reporte_asistencia(request):
    mes = request.GET.get("mes")
    anio = request.GET.get("anio")

    if not mes or not anio:
        return HttpResponse("Debes seleccionar un mes y un año.", status=400)

    mes = int(mes)
    anio = int(anio)

    wb = Workbook()
    wb.remove(wb.active)
    colaboradores = Collaborator.objects.enables()
    encabezados = ['Fecha', 'Ingreso', 'Salida', 'Refrigerio y/o almuerzo', 'Hora Total']
    peru_tz = pytz.timezone('America/Lima')
    
    for colaborador in colaboradores:
        hoja = wb.create_sheet(title=colaborador.user.first_name)
        
        hoja.merge_cells('A1:E1')
        hoja['A1'] = f'{datetime(anio, mes, 1).strftime("%B").upper()} {anio}'
        hoja['A1'].font = Font(bold=True)
        hoja['A1'].alignment = Alignment(horizontal='center')
        
        for col_num, encabezado in enumerate(encabezados, 1):
            hoja[get_column_letter(col_num) + '2'] = encabezado
        
        bold_font = Font(bold=True)
        for cell in hoja["2:2"]:
            cell.font = bold_font
        
        records = AttendanceRecord.objects.filter(
            collaborator=colaborador,
            isDelete=False,
            inRecord__dateTime__year=anio,
            inRecord__dateTime__month=mes
        ).order_by('inRecord__dateTime')
        
        fila = 3
        for record in records:
            fecha_entrada = localtime(record.inRecord.dateTime, peru_tz)
            hora_entrada = fecha_entrada.strftime('%H:%M')
            
            if record.outRecord:
                fecha_salida = localtime(record.outRecord.dateTime, peru_tz)
                hora_salida = fecha_salida.strftime('%H:%M')
                horas_trabajadas, launch = calculateHoursWorked(record)
            else:
                hora_salida = "0:00"
                horas_trabajadas = "0:00"
            
            hoja[f'A{fila}'] = fecha_entrada.strftime('%d/%m/%Y')
            hoja[f'B{fila}'] = hora_entrada
            hoja[f'C{fila}'] = hora_salida
            hoja[f'D{fila}'] = "1:00:00" if launch else "0:00:00"
            hoja[f'E{fila}'] = str(float_to_hms(horas_trabajadas)) if record.outRecord else "0:00:00"
            
            fila += 1
        
        # Ajuste de columnas
        for col in range(1, 6):
            hoja.column_dimensions[get_column_letter(col)].width = 20
        
        # Centrar el texto
        for row in hoja.iter_rows(min_row=1, max_row=fila - 1, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Reporte_Asistencia_{mes}_{anio}.xlsx'
    wb.save(response)
    return response